import PyPDF2
import re
import openpyxl
import sys
import tkinter as tk
from tkinter import messagebox

def merge_cells_if_needed(sheet, row, column):
    cell = sheet.cell(row=row, column=column)
    ranges_to_check = list(sheet.merged_cells.ranges)  # Create a list to iterate over
    for mergedCell in ranges_to_check:
        if (cell.coordinate not in mergedCell):
            NEIGHBOR = column +1
            sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=NEIGHBOR)

def process_stability(stability_file_path, manifest_file_path):
    ContainerNmb = []
    ContainerWeight = []
    ContainerSize = []
    sizeRegex = re.compile(r'(20|40|22)(?:DV|HC|RF|RH|TK|HW|K2|T6|G2)')
    weightRegex = re.compile(r'(?<!\d)\d{1,2}\.\d{3}')
    containerNumberRegex = re.compile(r'([A-Z]\s?[A-Z]\s?[A-Z]\s?[A-Z]\s?\d{6}\s\d)')
    pdfFileObj = open(manifest_file_path, 'rb')
    reader = PyPDF2.PdfReader(pdfFileObj)
    number_of_pages = len(reader.pages)

    for i in range(number_of_pages):
        page = reader.pages[i]
        text = page.extract_text()
        ContainerNmb.extend(containerNumberRegex.findall(text))
        ContainerWeight.extend(weightRegex.findall(text))
        ContainerSize.extend(sizeRegex.findall(text))


    ContainerDict = dict(zip(ContainerNmb, ContainerWeight))    # Your dictionary containing container numbers and weight values
    SizeDict = dict(zip(ContainerNmb, ContainerSize))
    ContainerCount = int(len(set(ContainerDict)))

    response = messagebox.askquestion('Verifieer container aantal', 'Het totaal aantal containers is ' + str(ContainerCount) + '.')
    if response == "yes":
        print("Container aantal is correct.")
    else:
        print("Container aantal is niet correct.")
        exit()
    #print(ContainerDict)
    #print(SizeDict)

    # Load the stability calculation Excel workbook
    stability_workbook = openpyxl.load_workbook(stability_file_path)
    stability_sheet = stability_workbook.active

    # Load the loading plan Excel workbook
    loading_plan_workbook = openpyxl.load_workbook(r'C:\Users\wgrin\Desktop\Stabiliteit\laadplanProgramma.xlsx')
    loading_plan_sheet = loading_plan_workbook.active

    # Define the range of rows and columns for the stability calculation
    stability_start_row = 3
    stability_end_row = 10
    stability_start_col = 2
    stability_end_col = 9


    # Iterate through the keys in the dictionary
    for container_number, weight_value in ContainerDict.items():
        found = False
        size = SizeDict.get(container_number, None)  # Get the container size from the dictionary

        # Iterate through cells in the stability calculation sheet within the specified range
        for row_index in range(stability_start_row, stability_end_row + 1):
            for col_index in range(stability_start_col, stability_end_col + 1):
                cell = stability_sheet.cell(row=row_index, column=col_index)
                if cell.value == float(weight_value):
                    # Search for a matching cell in the loading plan sheet
                    loading_plan_cell = loading_plan_sheet.cell(row=row_index, column=col_index)
                    if loading_plan_cell.value is None:
                        if size == '40':
                            merge_cells_if_needed(loading_plan_sheet, row_index, col_index)
                            loading_plan_cell.value = container_number
                        elif size == '20' or '22':
                            loading_plan_cell.value = container_number
                        else:
                            print('Containerformaat onbekend.')
                        found = True
                        break
            if found:
                break

        if not found:
            print(f"Gewicht {weight_value} niet gevonden in de stabiliteitsberekening.")

    # Save the modified loading plan sheet
    return loading_plan_workbook
